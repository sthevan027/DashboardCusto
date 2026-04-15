#!/usr/bin/env python3
"""
Gera Supabase/seed.generated.sql a partir do Excel de controle operacional (aba Dados).

Layout esperado (Controle Operacional V3.xlsx e equivalentes):
  A=Itens, B=Descrição, C=Sub-Grupo, D=UNID., E=QUANT., F=Valor Unid., G=Valor total,
  H=Total Com BDI, I=VALOR REAL.

Requer: pip install openpyxl

Regras:
- Linhas pai (coluna A = código): Total previsto/real -> grupo Total, item = descrição do pai.
- Filhas: col. B pode ser seção (MÃO DE OBRA / EQUIPAMENTOS / MATERIAIS) ou detalhe (ex.: tipo de
  equipamento); col. C = Sub-Grupo. Subtotais B=C=EQUIPAMENTOS ou MATERIAIS/MATERIAIS são ignorados
  na soma, mas atualizam a seção ativa para as linhas seguintes.
- Nome no banco: "{descrição pai} — {subgrupo}" (único por grupo; sufixo se colidir).
- Totais na aba Controle podem diferir da soma em Dados; o seed segue a aba Dados.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None


def get_openpyxl() -> Any:
    if openpyxl is None:
        print("Instale openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    return openpyxl


REPO = Path(__file__).resolve().parents[1]
EXCEL = REPO / "Excel" / "Controle Operacional V3.xlsx"
OUT = REPO / "Supabase" / "seed.generated.sql"
DEFAULT_SHEET = "Dados"

GROUP_CODE_PREFIX: dict[str, str] = {
    "Mão de Obra": "MO",
    "Equipamento": "EQ",
    "Materiais": "MAT",
    "Fornecimento": "FOR",
}


def sql_str(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"


def money(x: float) -> str:
    v = round(float(x) + 1e-9, 2)
    if v < 0:
        v = 0.0
    return f"{v:.2f}"


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_float(value: Any) -> float:
    try:
        return float(value) if value is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def norm_group(desc: str | None) -> str | None:
    """Só trata como seção MO/EQ/MAT linhas que são rótulos (não descrições longas)."""
    if not desc:
        return None
    t = str(desc).strip()
    u = t.upper()
    if u.startswith("MÃO") and "OBRA" in u:
        return "Mão de Obra"
    if u.startswith("EQUIPAMENTO"):
        return "Equipamento"
    if u.startswith("MATERIAIS") or u == "MATERIAL":
        return "Materiais"
    if u.startswith("FORNECIMENTO"):
        return "Fornecimento"
    return None


def normalize_equip_subgroup_name(subgroup_name: str) -> str:
    """Unifica singular/plural/typo ('Equipamento', 'Equipamentos', 'EquipamentoS') no mesmo subgrupo."""
    s = subgroup_name.strip()
    if not s or s == "—":
        return s
    nfd = unicodedata.normalize("NFD", s)
    folded = "".join(c for c in nfd if unicodedata.category(c) != "Mn").lower()
    compact = re.sub(r"\s+", "", folded)
    if compact in ("munk", "munck"):
        return "Munck"
    if compact in ("equipamento", "equipamentos"):
        return "Equipamentos (diversos)"
    return s


def normalize_subgroup(group: str, sub: str) -> str:
    """Unifica grafias de 'Mão de Obra' e nomes genéricos duplicados em Equipamento."""
    s = (sub or "—").strip()
    if not s:
        s = "—"
    if group == "Mão de Obra":
        folded = (
            unicodedata.normalize("NFKD", s)
            .encode("ascii", "ignore")
            .decode()
            .lower()
        )
        if "mao" in folded and "de" in folded and "obra" in folded:
            return "Mão de Obra"
        return s
    if group == "Equipamento":
        return normalize_equip_subgroup_name(s)
    return s


def is_section_subtotal(b: str, c: str) -> bool:
    """Subtotal quando B e C são a mesma célula (ex.: EQUIPAMENTOS|EQUIPAMENTOS), não B=EQUIPAMENTOS C=Equipamentos."""
    b0 = str(b).strip()
    c0 = str(c).strip()
    if b0 != c0:
        return False
    u = b0.upper()
    return u in ("EQUIPAMENTOS", "MATERIAIS", "FORNECIMENTOS")


def slug_code(prefix: str, name: str, used: dict[str, int]) -> str:
    raw = re.sub(r"[^A-Za-z0-9]+", "-", name.strip())
    raw = re.sub(r"-+", "-", raw).strip("-").upper()
    if not raw:
        raw = "X"
    base = f"{prefix}-{raw}"[:55]
    if base not in used:
        used[base] = 1
        return base
    n = used[base]
    used[base] = n + 1
    return f"{base}-{n}"[:60]


def dedupe_redundant_mat_materiais_header(parents: list[dict]) -> None:
    """
    Quando o Excel traz uma linha MAT 'Materiais' (resumo) seguida de detalhes
    (Andaime, Outros, …) que somam o mesmo valor, remove o resumo para não
    duplicar o orçamento.
    """
    for p in parents:
        lines = p["breakdown"]
        drop: list[int] = []
        for i, b in enumerate(lines):
            if b["group"] != "Materiais":
                continue
            if b["subgroup"].strip().lower() != "materiais":
                continue
            sub_sum = 0.0
            for j in range(i + 1, len(lines)):
                nxt = lines[j]
                if nxt["group"] != "Materiais":
                    break
                if nxt["subgroup"].strip().lower() == "materiais":
                    break
                sub_sum += nxt["planned"]
            if sub_sum > 0 and abs(sub_sum - b["planned"]) < 0.02:
                drop.append(i)
        for i in reversed(drop):
            del lines[i]
        recompute_display_labels(lines)


def recompute_display_labels(lines: list[dict]) -> None:
    order: dict[tuple[str, str], int] = defaultdict(int)
    for b in lines:
        key = (b["group"], b["subgroup"])
        order[key] += 1
        n = order[key]
        sg = b["subgroup"]
        b["display_label"] = sg if n == 1 else f"{sg} (#{n})"


def find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Localiza a linha cujo texto em A é 'Itens' (cabeçalho da tabela de dados)."""
    for i, row in enumerate(rows[:30]):
        a = as_text(row[0] if row else None).lower()
        if a == "itens":
            return i
    return 3


def row_planned_and_real(row: tuple[Any, ...]) -> tuple[float, float]:
    """
    Orçamento (previsto) = coluna Valor total (G); se vazio, QUANT * Valor Unid.
    Real = coluna VALOR REAL (I), ou 0.
    """
    q = as_float(row[4] if len(row) > 4 else None)
    vu = as_float(row[5] if len(row) > 5 else None)
    vt = row[6] if len(row) > 6 else None
    if vt is not None and str(vt).strip() != "":
        tp = as_float(vt)
    else:
        tp = round(q * vu + 1e-9, 2)
    tr_raw = row[8] if len(row) > 8 else None
    tr = as_float(tr_raw) if tr_raw is not None else 0.0
    return tp, tr


def parse_dados(excel_path: Path, sheet_name: str) -> tuple[list[dict], dict[str, float]]:
    openpyxl_mod = get_openpyxl()
    wb = openpyxl_mod.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        wb.close()
        raise KeyError(
            f"Aba '{sheet_name}' não encontrada em {excel_path.name}. Abas disponíveis: {available}"
        )
    ws = wb[sheet_name]
    rows: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    wb.close()

    header_i = find_header_row(rows)
    data_start = header_i + 1

    parents: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    def push_breakdown(group: str, subgroup: str, tp: float, tr: float) -> None:
        assert current is not None
        sg = normalize_subgroup(group, subgroup)
        key = (group, sg)
        current["sub_order"][key] += 1
        n = current["sub_order"][key]
        display = sg if n == 1 else f"{sg} (#{n})"
        current["breakdown"].append(
            {
                "group": group,
                "subgroup": sg,
                "display_label": display,
                "planned": tp,
                "real": tr,
            }
        )

    for row in rows[data_start:]:
        if not row:
            continue
        code = as_text(row[0])
        desc = as_text(row[1])
        sub = as_text(row[2])
        tp, tr = row_planned_and_real(row)

        if code and re.match(r"^[\d.]+", code):
            if current:
                parents.append(current)
            current = {
                "code": code,
                "name": desc,
                "subgrp_header": sub,
                "planned": tp,
                "real": tr,
                "breakdown": [],
                "section": None,
                "sub_order": defaultdict(int),
            }
        elif current:
            b_raw = as_text(row[1] if len(row) > 1 else None)
            c_raw = as_text(row[2] if len(row) > 2 else None)
            if not b_raw and not c_raw:
                continue

            if is_section_subtotal(b_raw, c_raw):
                head = b_raw.strip().upper()
                if head == "EQUIPAMENTOS":
                    current["section"] = "Equipamento"
                elif head == "MATERIAIS":
                    current["section"] = "Materiais"
                elif head == "FORNECIMENTOS":
                    current["section"] = "Fornecimento"
                continue

            nb = norm_group(b_raw)
            if nb:
                current["section"] = nb
                push_breakdown(nb, c_raw or b_raw, tp, tr)
            else:
                sec = current.get("section")
                if not isinstance(sec, str) or not sec:
                    continue
                leaf_sub = (c_raw or b_raw or "—").strip() or "—"
                push_breakdown(sec, leaf_sub, tp, tr)

    if current:
        parents.append(current)

    dedupe_redundant_mat_materiais_header(parents)

    sums = {
        "parent_planned": sum(p["planned"] for p in parents),
        "parent_real": sum(p["real"] for p in parents),
        "bd_planned": 0.0,
        "bd_real": 0.0,
    }
    for p in parents:
        for b in p["breakdown"]:
            sums["bd_planned"] += b["planned"]
            sums["bd_real"] += b["real"]

    return parents, sums


def build_subgroups(parents: list[dict]) -> list[tuple[str, str, str]]:
    """(group_name, subgroup_name, code) for INSERT."""
    used: dict[str, int] = {}
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str, str]] = []

    def add(gn: str, sn: str, pfx: str) -> None:
        if (gn, sn) in seen:
            return
        seen.add((gn, sn))
        code = slug_code(pfx, sn, used)
        out.append((gn, sn, code))

    add("Total", "Total", "TOT")
    add("Mão de Obra", "Mão de Obra", "MO")

    for p in parents:
        for b in p["breakdown"]:
            gn = b["group"]
            sn = b["subgroup"]
            pfx = GROUP_CODE_PREFIX.get(gn)
            if not pfx:
                raise ValueError(
                    f"Grupo não mapeado em GROUP_CODE_PREFIX: {gn!r}"
                )
            add(gn, sn, pfx)

    out.sort(key=lambda x: (x[0], x[1]))
    return out


def item_breakdown_name(parent_name: str, display_label: str) -> str:
    return f"{parent_name} — {display_label}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera o seed SQL do Supabase a partir de uma aba do Excel."
    )
    parser.add_argument(
        "--excel",
        default=str(EXCEL),
        help="Caminho do arquivo Excel. Padrão: Excel/Controle Operacional V3.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Nome da aba a ser lida. Padrão: {DEFAULT_SHEET}",
    )
    parser.add_argument(
        "--out",
        default=str(OUT),
        help="Arquivo SQL de saída. Padrão: Supabase/seed.generated.sql",
    )
    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="Lista as abas disponíveis no Excel informado e encerra.",
    )
    return parser.parse_args()


def resolve_user_path(value: str) -> Path:
    path = Path(value).expanduser()
    if path.is_absolute():
        return path
    return (Path.cwd() / path).resolve()


def generate() -> int:
    args = parse_args()
    excel_path = resolve_user_path(args.excel)
    out_path = resolve_user_path(args.out)

    if not excel_path.is_file():
        print(f"Arquivo não encontrado: {excel_path}", file=sys.stderr)
        return 1

    if args.list_sheets:
        openpyxl_mod = get_openpyxl()
        wb = openpyxl_mod.load_workbook(excel_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            print(sheet_name)
        wb.close()
        return 0

    try:
        parents, sums = parse_dados(excel_path, args.sheet)
        subgroups = build_subgroups(parents)
    except (KeyError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 1

    lines: list[str] = []
    w = lines.append

    w("-- Seed gerado automaticamente (scripts/generate_seed_from_excel.py)")
    source_label = excel_path.relative_to(REPO) if excel_path.is_relative_to(REPO) else excel_path
    w(f"-- Fonte: {source_label} — aba {args.sheet}")
    w("--")
    w("-- Grupo Total: Total previsto/real por código (linha pai).")
    w("-- Grupos MO/EQ/MAT: quebra por subgrupo; nome do item = \"descrição — subgrupo\".")
    w("--")
    w(f"-- Nota: a soma dos Total previsto das linhas pai em {args.sheet} pode diferir do")
    w("-- \"Total\" na aba Controle (~8.04M vs ~8.09M). O seed reflete Dados.")
    w("--")
    w(f"-- Sanity (aba {args.sheet}):")
    w(f"--   Soma Total previsto (linhas pai): {sums['parent_planned']:.2f}")
    w(f"--   Soma Total real (linhas pai): {sums['parent_real']:.2f}")
    w(f"--   Soma previsto (MO+EQ+MAT detalhado): {sums['bd_planned']:.2f}")
    w(f"--   Soma real (MO+EQ+MAT detalhado): {sums['bd_real']:.2f}")
    w("")

    w("-- Groups")
    w("insert into public.cost_groups (name, code) values")
    w("  ('Mão de Obra','MO'),")
    w("  ('Equipamento','EQ'),")
    w("  ('Materiais','MAT'),")
    w("  ('Fornecimento','FOR'),")
    w("  ('Total','TOT')")
    w("on conflict (name) do nothing;")
    w("")

    w("-- Subgroups")
    w("insert into public.cost_subgroups (group_id, name, code)")
    w("select g.id, s.name, s.code")
    w("from public.cost_groups g")
    w("join (values")
    for i, (gn, sn, sc) in enumerate(subgroups):
        comma = "," if i < len(subgroups) - 1 else ""
        w(f"  ({sql_str(gn)}, {sql_str(sn)}, {sql_str(sc)}){comma}")
    w(") as s(group_name, name, code) on s.group_name=g.name")
    w("on conflict (group_id, name) do nothing;")
    w("")

    w("-- Items (totais por código)")
    w("insert into public.cost_items (group_id, subgroup_id, name, code)")
    w("select g.id, sg.id, i.name, i.code")
    w("from public.cost_groups g")
    w("join public.cost_subgroups sg on sg.group_id=g.id")
    w("join (values")
    for i, p in enumerate(parents):
        comma = "," if i < len(parents) - 1 else ""
        w(
            f"  ('Total', 'Total', {sql_str(p['name'])}, {sql_str(p['code'])}){comma}"
        )
    w(") as i(group_name, subgroup_name, name, code)")
    w("  on i.group_name=g.name and i.subgroup_name=sg.name")
    w("on conflict (group_id, name) do nothing;")
    w("")

    w("-- Items (quebra por grupo/subgrupo quando existir)")
    w("insert into public.cost_items (group_id, subgroup_id, name, code)")
    w("select g.id, sg.id, i.name, i.code")
    w("from public.cost_groups g")
    w("join public.cost_subgroups sg on sg.group_id=g.id")
    w("join (values")
    bd_rows: list[tuple[str, str, str, str]] = []
    for p in parents:
        for b in p["breakdown"]:
            nm = item_breakdown_name(p["name"], b["display_label"])
            bd_rows.append((b["group"], b["subgroup"], nm, p["code"]))
    for i, row in enumerate(bd_rows):
        gn, sn, nm, code = row
        comma = "," if i < len(bd_rows) - 1 else ""
        w(f"  ({sql_str(gn)}, {sql_str(sn)}, {sql_str(nm)}, {sql_str(code)}){comma}")
    w(") as i(group_name, subgroup_name, name, code)")
    w("  on i.group_name=g.name and i.subgroup_name=sg.name")
    w("on conflict (group_id, name) do nothing;")
    w("")

    w("-- Budgets (total por código)")
    w("insert into public.cost_budgets (item_id, planned_value, currency_code, effective_date)")
    w("select it.id, b.planned_value, 'BRL', date '2026-01-01'")
    w("from public.cost_items it")
    w("join public.cost_groups g on g.id=it.group_id")
    w("join public.cost_subgroups sg on sg.id=it.subgroup_id")
    w("join (values")
    for i, p in enumerate(parents):
        comma = "," if i < len(parents) - 1 else ""
        w(
            f"  ('Total', 'Total', {sql_str(p['name'])}, {sql_str(p['code'])}, {money(p['planned'])}){comma}"
        )
    w(") as b(group_name, subgroup_name, item_name, item_code, planned_value)")
    w("  on b.group_name=g.name and b.subgroup_name=sg.name")
    w("    and b.item_name=it.name and b.item_code=it.code")
    w("on conflict (item_id) do nothing;")
    w("")

    w("-- Budgets (quebra por grupo/subgrupo)")
    w("insert into public.cost_budgets (item_id, planned_value, currency_code, effective_date)")
    w("select it.id, b.planned_value, 'BRL', date '2026-01-01'")
    w("from public.cost_items it")
    w("join public.cost_groups g on g.id=it.group_id")
    w("join public.cost_subgroups sg on sg.id=it.subgroup_id")
    w("join (values")
    budget_break_rows: list[tuple[str, str, str, str, str]] = []
    for p in parents:
        for b in p["breakdown"]:
            nm = item_breakdown_name(p["name"], b["display_label"])
            budget_break_rows.append(
                (b["group"], b["subgroup"], nm, p["code"], money(b["planned"]))
            )
    for i, (gn, sn, nm, code, pv) in enumerate(budget_break_rows):
        comma = "," if i < len(budget_break_rows) - 1 else ""
        w(
            f"  ({sql_str(gn)}, {sql_str(sn)}, {sql_str(nm)}, {sql_str(code)}, {pv}){comma}"
        )
    w(") as b(group_name, subgroup_name, item_name, item_code, planned_value)")
    w("  on b.group_name=g.name and b.subgroup_name=sg.name")
    w("    and b.item_name=it.name and b.item_code=it.code")
    w("on conflict (item_id) do nothing;")
    w("")

    desc = f"Posição atual ({excel_path.name} — aba {args.sheet})"

    w("-- Costs (real total por código)")
    w("insert into public.cost_entries (item_id, cost_date, amount, description, external_id)")
    w("select it.id, date '2026-04-01', c.amount, c.description, c.external_id")
    w("from public.cost_items it")
    w("join public.cost_groups g on g.id=it.group_id")
    w("join public.cost_subgroups sg on sg.id=it.subgroup_id")
    w("join (values")
    for i, p in enumerate(parents):
        comma = "," if i < len(parents) - 1 else ""
        ext = f"SEED-V2-TOT-{p['code'].replace('.', '-')}"
        w(
            f"  ('Total', 'Total', {sql_str(p['name'])}, {sql_str(p['code'])}, {money(p['real'])}, {sql_str(desc)}, {sql_str(ext)}){comma}"
        )
    w(") as c(group_name, subgroup_name, item_name, item_code, amount, description, external_id)")
    w("  on c.group_name=g.name and c.subgroup_name=sg.name")
    w("    and c.item_name=it.name and c.item_code=it.code")
    w("on conflict (item_id, external_id) do nothing;")
    w("")

    w("-- Costs (real por grupo/subgrupo)")
    w("insert into public.cost_entries (item_id, cost_date, amount, description, external_id)")
    w("select it.id, date '2026-04-01', c.amount, c.description, c.external_id")
    w("from public.cost_items it")
    w("join public.cost_groups g on g.id=it.group_id")
    w("join public.cost_subgroups sg on sg.id=it.subgroup_id")
    w("join (values")
    ext_i = 0
    rows = []
    for p in parents:
        for b in p["breakdown"]:
            nm = item_breakdown_name(p["name"], b["display_label"])
            pfx = GROUP_CODE_PREFIX[b["group"]]
            ext = f"SEED-V2-{pfx}-{p['code'].replace('.', '-')}-{ext_i}"
            ext_i += 1
            rows.append(
                (
                    b["group"],
                    b["subgroup"],
                    nm,
                    p["code"],
                    b["real"],
                    desc,
                    ext,
                )
            )
    for i, r in enumerate(rows):
        gn, sn, nm, code, amt, d, ext = r
        comma = "," if i < len(rows) - 1 else ""
        w(
            f"  ({sql_str(gn)}, {sql_str(sn)}, {sql_str(nm)}, {sql_str(code)}, {money(amt)}, {sql_str(d)}, {sql_str(ext)}){comma}"
        )
    w(") as c(group_name, subgroup_name, item_name, item_code, amount, description, external_id)")
    w("  on c.group_name=g.name and c.subgroup_name=sg.name")
    w("    and c.item_name=it.name and c.item_code=it.code")
    w("on conflict (item_id, external_id) do nothing;")
    w("")

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(
        f"Escrito: {out_path} ({len(parents)} itens pai, {sum(len(p['breakdown']) for p in parents)} linhas de quebra)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(generate())
